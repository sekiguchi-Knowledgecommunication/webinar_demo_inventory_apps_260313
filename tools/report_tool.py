"""
レポート自動生成ツール — 分析結果を Excel ファイルとして保存

エージェントが分析結果を構造化レポートとしてまとめ、
リッチな Excel ファイルに保存してダウンロード用パスを返す。
"""

import os
import datetime
import logging
import io

logger = logging.getLogger(__name__)

# レポート保存ディレクトリ
REPORT_DIR = "/Workspace/Users/s.sekiguchi7056@gmail.com/10.webinar/00.260313/webinar_demo"

# デザイン定数（Databricks ブランドカラー）
COLOR_HEADER_BG    = "1B3A6B"  # ダークネイビー（ヘッダー背景）
COLOR_HEADER_FG    = "FFFFFF"  # 白文字
COLOR_TITLE_BG     = "FF3621"  # Databricks レッド（タイトル帯）
COLOR_TITLE_FG     = "FFFFFF"  # 白文字
COLOR_ROW_ODD      = "F0F4FF"  # 淡ブルー（奇数行）
COLOR_ROW_EVEN     = "FFFFFF"  # 白（偶数行）
COLOR_META_FG      = "6B7280"  # グレー（メタ情報）
COLOR_CRITICAL_BG  = "FFF0F0"  # 薄赤（危険ラインの強調）
COLOR_CRITICAL_FG  = "DC2626"  # 赤文字


from agents import function_tool

@function_tool
def generate_report(
    report_title: str,
    headers: list,
    rows: list,
    summary: str,
) -> str:
    """
    分析結果をリッチな Excel レポートとして保存する。

    エージェントが Genie で取得したデータを整理し、
    ユーザーがダウンロードできる Excel ファイルとして出力する。
    タイトル行・ヘッダー行のカラー、交互背景色（ゼブラストライプ）、
    数値フォーマットを適用した視覚的に分かりやすいレポートを生成する。

    Args:
        report_title: レポートのタイトル（例: \"過剰在庫分析レポート\"）
        headers: 列ヘッダー行（例: [\"品目ID\", \"品目名\", \"在庫金額\"]）
        rows: データ行のリスト（例: [[\"ITM-001\", \"電子部品A\", \"1200000\"], ...]）
        summary: レポートのサマリテキスト

    Returns:
        レポートファイルのパスと概要を含むメッセージ
    """
    try:
        # ファイル名の生成（.xlsx 拡張子）
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_title = report_title.replace(" ", "_").replace("/", "_").replace("\\", "_")[:30]
        filename  = f"{safe_title}_{timestamp}.xlsx"
        filepath  = os.path.join(REPORT_DIR, filename)

        # Excel ファイルをメモリ上に生成
        excel_bytes = _build_excel(report_title, headers, rows, summary)

        # Databricks SDK 経由で Workspace にアップロード
        from databricks.sdk import WorkspaceClient
        from databricks.sdk.service.workspace import ImportFormat
        import base64

        try:
            w = WorkspaceClient()
            try:
                w.workspace.mkdirs(REPORT_DIR)
            except Exception as d_err:
                logger.warning(f"ディレクトリ作成で例外が発生しましたが続行します: {d_err}")

            b64_str = base64.b64encode(excel_bytes).decode("utf-8")
            w.workspace.import_(
                path=filepath,
                content=b64_str,
                format=ImportFormat.AUTO,
                overwrite=True,
            )
        except Exception as upload_err:
            logger.error(f"Workspace へのアップロードに失敗しました: {upload_err}")
            return f"⚠️ レポートの保存に失敗しました: {upload_err}"

        row_count = len(rows)
        logger.info(f"📊 Excel レポート生成完了: {filepath} ({row_count} 行)")

        return (
            f"[REPORT:{filepath}]\n\n"
            f"📊 **Excel レポートを生成しました**\n\n"
            f"- **タイトル**: {report_title}\n"
            f"- **データ件数**: {row_count} 件\n"
            f"- **ファイル**: {filename}\n\n"
            f"**サマリ**: {summary}"
        )

    except Exception as e:
        logger.error(f"レポート生成エラー: {e}")
        import traceback
        traceback.print_exc()
        return f"⚠️ レポートの生成中にエラーが発生しました: {e}"


def _build_excel(
    report_title: str,
    headers: list,
    rows: list,
    summary: str,
) -> bytes:
    """
    openpyxl を使ってリッチな Excel ファイルを構築し、bytes で返す。

    シート構成:
        Sheet 1「データ」  — タイトル・メタ情報・ヘッダー・データ行・合計行
        Sheet 2「サマリ」  — AI エージェントのサマリテキスト
    """
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise RuntimeError(
            "openpyxl がインストールされていません。requirements.txt に openpyxl>=3.1.0 を追加してください。"
        ) from e

    wb = openpyxl.Workbook()

    # ============================
    # Sheet 1: データシート
    # ============================
    ws = wb.active
    ws.title = "データ"

    # --- スタイル定義 ---
    font_title   = Font(name="Meiryo UI", bold=True, size=14, color=COLOR_TITLE_FG)
    font_header  = Font(name="Meiryo UI", bold=True, size=10, color=COLOR_HEADER_FG)
    font_meta    = Font(name="Meiryo UI", size=9,  color=COLOR_META_FG)
    font_body    = Font(name="Meiryo UI", size=10)
    font_total   = Font(name="Meiryo UI", bold=True, size=10)
    font_critical = Font(name="Meiryo UI", bold=True, size=10, color=COLOR_CRITICAL_FG)

    fill_title   = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    fill_header  = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    fill_odd     = PatternFill("solid", fgColor=COLOR_ROW_ODD)
    fill_even    = PatternFill("solid", fgColor=COLOR_ROW_EVEN)
    fill_critical = PatternFill("solid", fgColor=COLOR_CRITICAL_BG)
    fill_total   = PatternFill("solid", fgColor="E8EDF5")

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    align_right  = Alignment(horizontal="right",  vertical="center")

    thin_side = Side(style="thin", color="D1D5DB")
    border_header = Border(
        bottom=Side(style="medium", color=COLOR_HEADER_BG),
        left=thin_side, right=thin_side,
    )
    border_data = Border(
        left=thin_side, right=thin_side,
        bottom=thin_side, top=thin_side,
    )

    col_count = max(len(headers), 1) if headers else 1

    # --- 行1: タイトル ---
    ws.row_dimensions[1].height = 32
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
    title_cell = ws.cell(row=1, column=1, value=f"📊 {report_title}")
    title_cell.font      = font_title
    title_cell.fill      = fill_title
    title_cell.alignment = align_center

    # --- 行2: 作成日時 ---
    ws.row_dimensions[2].height = 18
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
    meta_str = (
        f"作成日時: {datetime.datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}   "
        f"作成者: AI エージェント（Databricks Genie）"
    )
    meta_cell = ws.cell(row=2, column=1, value=meta_str)
    meta_cell.font      = font_meta
    meta_cell.alignment = align_right

    # --- 行3: 空行 ---
    ws.row_dimensions[3].height = 6

    # --- 行4: ヘッダー ---
    HEADER_ROW = 4
    ws.row_dimensions[HEADER_ROW].height = 22
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=header_text)
        cell.font      = font_header
        cell.fill      = fill_header
        cell.alignment = align_center
        cell.border    = border_header

    # --- 行5以降: データ行 ---
    # 数値っぽいカラムを判定するキーワード
    NUMERIC_KEYWORDS = ("金額", "数量", "qty", "value", "rate", "率", "日数", "days", "price", "amount", "pct", "%")
    CURRENCY_KEYWORDS = ("金額", "value", "price", "amount")

    numeric_cols: set[int] = set()
    currency_cols: set[int] = set()
    for col_idx, h in enumerate(headers, start=1):
        h_lower = h.lower()
        if any(kw in h_lower for kw in NUMERIC_KEYWORDS):
            numeric_cols.add(col_idx)
        if any(kw in h_lower for kw in CURRENCY_KEYWORDS):
            currency_cols.add(col_idx)

    DATA_START_ROW = HEADER_ROW + 1
    for row_idx, row_data in enumerate(rows):
        excel_row = DATA_START_ROW + row_idx
        ws.row_dimensions[excel_row].height = 18
        fill = fill_odd if row_idx % 2 == 0 else fill_even

        # "危険ライン" の強調: 回転率が1.5未満、または在庫日数が300以上の場合
        is_critical = False
        for col_idx, cell_val in enumerate(row_data, start=1):
            h_lower = headers[col_idx - 1].lower() if col_idx <= len(headers) else ""
            try:
                num_val = float(str(cell_val).replace(",", "").replace("¥", "").replace("%", ""))
                if ("回転率" in h_lower or "turnover" in h_lower) and num_val < 1.5:
                    is_critical = True
                if ("日数" in h_lower or "days" in h_lower) and num_val >= 300:
                    is_critical = True
            except (ValueError, TypeError):
                pass

        for col_idx, cell_val in enumerate(row_data, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.border = border_data

            # 数値変換を試みる
            converted_val = _try_convert_number(cell_val)
            cell.value = converted_val

            if is_critical:
                cell.fill  = fill_critical
                cell.font  = font_critical
            else:
                cell.fill = fill
                cell.font = font_body

            # 数値列の書式
            if col_idx in currency_cols and isinstance(converted_val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = align_right
            elif col_idx in numeric_cols and isinstance(converted_val, (int, float)):
                cell.number_format = '#,##0.##'
                cell.alignment = align_right
            else:
                cell.alignment = align_left

    # --- 合計行（数値列のみ SUM） ---
    if rows:
        total_row = DATA_START_ROW + len(rows)
        ws.row_dimensions[total_row].height = 20
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=min(2, col_count))
        total_label = ws.cell(row=total_row, column=1, value="合 計")
        total_label.font      = font_total
        total_label.fill      = fill_total
        total_label.alignment = align_center
        total_label.border    = border_data

        for col_idx in range(3 if col_count > 2 else 2, col_count + 1):
            cell = ws.cell(row=total_row, column=col_idx)
            cell.font   = font_total
            cell.fill   = fill_total
            cell.border = border_data

            if col_idx in currency_cols or col_idx in numeric_cols:
                # SUM 数式（数値列のみ）
                col_letter  = get_column_letter(col_idx)
                data_start  = get_column_letter(col_idx) + str(DATA_START_ROW)
                data_end    = get_column_letter(col_idx) + str(DATA_START_ROW + len(rows) - 1)
                cell.value  = f"=SUM({data_start}:{data_end})"
                if col_idx in currency_cols:
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0.##'
                cell.alignment = align_right
            else:
                cell.value = "—"
                cell.alignment = align_center

    # --- 列幅の自動調整 ---
    for col_idx in range(1, col_count + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1])) if col_idx <= len(headers) else 8
        for row_data in rows:
            if col_idx <= len(row_data):
                cell_str = str(row_data[col_idx - 1])
                max_len = max(max_len, len(cell_str))
        # 日本語は全角を考慮（概算: 文字数 × 1.8、最小10、最大40）
        adjusted = max(10, min(45, int(max_len * 1.8) + 2))
        ws.column_dimensions[col_letter].width = adjusted

    # 先頭列固定（スクロール時にも品目名が見えるように）
    ws.freeze_panes = f"C{HEADER_ROW + 1}"

    # ============================
    # Sheet 2: サマリシート
    # ============================
    ws2 = wb.create_sheet(title="サマリ")
    ws2.row_dimensions[1].height = 28
    ws2.column_dimensions["A"].width = 80

    ws2.merge_cells("A1:E1")
    hdr = ws2.cell(row=1, column=1, value="📋 AI エージェント 分析サマリ")
    hdr.font      = Font(name="Meiryo UI", bold=True, size=13, color=COLOR_HEADER_FG)
    hdr.fill      = fill_header
    hdr.alignment = align_center

    ws2.row_dimensions[2].height = 8
    ws2.merge_cells("A3:E3")
    date_cell = ws2.cell(row=3, column=1, value=f"作成日時: {datetime.datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}")
    date_cell.font = font_meta
    date_cell.alignment = align_right

    ws2.row_dimensions[4].height = 8
    ws2.row_dimensions[5].height = 14

    # サマリテキストを段落ごとに行に分割して記述
    summary_lines = summary.replace("\\n", "\n").split("\n")
    for line_idx, line in enumerate(summary_lines):
        row_num = 5 + line_idx
        ws2.row_dimensions[row_num].height = 15
        ws2.merge_cells(
            start_row=row_num, start_column=1,
            end_row=row_num, end_column=min(col_count, 5),
        )
        cell = ws2.cell(row=row_num, column=1, value=line)
        cell.font      = Font(name="Meiryo UI", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # サマリシートの列幅
    ws2.column_dimensions["A"].width = 90

    # ============================
    # バイト列として返す
    # ============================
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _try_convert_number(value):
    """文字列を数値（int/float）に変換を試みる。変換不可の場合は元の値を返す。"""
    if isinstance(value, (int, float)):
        return value
    if value is None:
        return ""
    s = str(value).strip().replace(",", "").replace("¥", "").replace("％", "")
    try:
        if "." in s:
            return float(s)
        return int(s)
    except (ValueError, TypeError):
        return str(value)
